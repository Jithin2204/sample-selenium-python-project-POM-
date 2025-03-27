from selenium import webdriver
import unittest
from openpyxl import load_workbook
import time

from webdriver_manager.chrome import ChromeDriverManager
from Pages.seachCorrectCityName import SearchCorrectCity
from Pages.seachIncorrectCityName import SearchIncorrectCity
from Assets.assets import Assets

class SearchCity(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.driver = webdriver.Chrome(ChromeDriverManager().install())

        cls.driver.implicitly_wait(15)
        cls.driver.maximize_window()
        cls.driver.get(Assets.weather_forecast_url)
        time.sleep(20)


    def test_01_search_valid_city(self):
        driver = self.driver

        self.workbook = load_workbook(Assets.workbook_path)
        self.workbook.active = 0
        self.sheet = self.workbook.active

        for i in range(1, self.sheet.max_row + 1):
            self.value_city_name = self.sheet.cell(row=i, column=1).value
            self.value_city_name_str = str(self.value_city_name)

            # Search for a valid city
            self.driver.get(Assets.weather_forecast_url)
            search_city = SearchCorrectCity(driver)
            search_city.search_correct_city_name(self.value_city_name_str)

    def test_02_search_invalid_city(self):
        driver = self.driver
        self.workbook = load_workbook(Assets.workbook_path)
        self.workbook.active = 1
        self.sheet = self.workbook.active

        for i in range(1, self.sheet.max_row + 1):
            self.value_city_name = self.sheet.cell(row=i, column=1).value
            self.value_city_name_str = str(self.value_city_name)

            # Search for a valid city
            self.driver.get(Assets.weather_forecast_url)
            search_city = SearchIncorrectCity(driver)
            search_city.search_incorrect_city_name(self.value_city_name_str)

    @classmethod
    def tearDownClass(cls):
        print("Test completed successfully")
        cls.driver.close()
        cls.driver.quit()
